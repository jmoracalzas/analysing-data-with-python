#!/usr/bin/env python
from openpyxl import Workbook
from openpyxl import load_workbook


class TXTFiles:
    def __init__(self, ccList, incList, expDict, dataList):
        self.__ccList = ccList
        self.__incList = incList
        self.__expDict = expDict
        self.__userData = dataList
        self.__path = "./Project_1/python/output/txt_files/"

    # creating the txt file and producing the desired output
    def newFile(self, title, fileName):

        with open(self.__path + fileName, "tw") as outputFile:
            outputFile.write(title + "\n")
            outputFile.write(str((len(title) + 10) * "-") + "\n")
        return None

    # appending the data to the new files
    def appendSettings(self, fileName, list):
        with open(self.__path + fileName, "ta") as outputFile:
            if fileName == "costCentres.txt" or fileName == "income.txt":
                for item in list:
                    outputFile.write(item + "\n")

            elif fileName == "expenditure.txt":
                outputFile.write(
                    "Expenditure\tClassification\t%_Sales\t%_Distribution\t%_Production\t%Admin\tMax_Cost\n"
                )

                for item in list.items():
                    line = "{0}\t{1}\t{2}\t{3}\t{4}\t{5}\t{6}".format(
                        item[0],  # type
                        item[1][0],  # classification
                        item[1][1],  #%_Sales
                        item[1][2],  #%_Distribution
                        item[1][3],  #%_Production
                        item[1][4],  #%_Admin
                        item[1][5],  # Max_Cost
                    )
                    outputFile.write(line + "\n")

            elif fileName == "dataset.txt":
                outputFile.write(
                    "Period\tType\tCategory\tAccount\tCost Centre\tDescription\tAmount\n"
                )
                # print(list)
                for item in list:
                    outputFile.write(item + "\n")

        return None

    def createTXTfiles(self):
        self.incTypes()
        self.costCentre()
        self.expTypes()
        self.genDataSet()

        return None

    # exporting the cost centres as a .txt file
    def costCentre(self):
        self.newFile(title="Cost Centres", fileName="costCentres.txt")
        self.appendSettings(fileName="costCentres.txt", list=self.__ccList)

    # exporting the income types as a .txt file
    def incTypes(self):
        self.newFile(title="Income type", fileName="income.txt")
        self.appendSettings(fileName="income.txt", list=self.__incList)

    def expTypes(self):
        self.newFile(title="Expenditure type", fileName="expenditure.txt")
        self.appendSettings(fileName="expenditure.txt", list=self.__expDict)

    def genDataSet(self):
        self.newFile(title="Generated data", fileName="dataset.txt")
        self.appendSettings(fileName="dataset.txt", list=self.__userData)


class ExcelExport:
    def __init__(self, ccList, incList, userData):
        self.__path = "./Project_1/python/output/xls_files/"
        self.__ccList = ccList
        self.__incList = incList
        self.__userData = userData

    def createXLSX(self, file):
        # creating the file structure
        self.xlsStructure(file)

        # exporting the main settings
        self.settingsExp()

        # exporting the main dataset generated by the user
        self.userDataExp()

    def xlsStructure(self, file):
        # creating the dataset worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Dataset"

        # adding the header to the dataset worksheet
        ws["A1"] = "Date"
        ws["B1"] = "Transaction Type"
        ws["C1"] = "Nature"
        ws["D1"] = "Account"
        ws["E1"] = "Cost Centre"
        ws["F1"] = "Description"
        ws["G1"] = "Amount"

        # Creating the settings worksheet
        ws1 = wb.create_sheet("Cost Centres")
        ws1.title = "Settings"

        # accessing the "Settings" worksheet and adding the header
        wb.active = 1
        ws1["A1"] = "Cost_Centres"
        ws1["C1"] = "Income Categories"
        ws1["E1"] = "Expenditure Types"
        ws1["F1"] = "Classification"
        ws1["G1"] = "%_Sales"
        ws1["H1"] = "%_Distribution"
        ws1["I1"] = "%_Production"
        ws1["J1"] = "%_Admin"
        ws1["K1"] = "Max_Cost"

        wb.save(self.__path + "dataset.xlsx")

    def settingsExp(self):
        # exporting the list of cost centres
        self.singleColumnSettings(self.__ccList, uniqueColumn=1)

        # exporting the list of income categories
        self.singleColumnSettings(self.__incList, uniqueColumn=3)

    def userDataExp(self):
        # loading the file
        wb = load_workbook(self.__path + "dataset.xlsx")
        ws = wb["Dataset"]

        # obtaining the user data to transfer to the spreadsheet
        stackUserData = self.__userData[:]
        row_limit = len(stackUserData) + 1

        # exporting the data on to Microsoft Excel
        # transferring the data
        for row in ws.iter_rows(min_row=2, max_col=7, min_col=1, max_row=row_limit):
            # extracting the last element of the stack
            dataLine = stackUserData.pop()

            # separating the elements and revering the their order
            elements = dataLine.split("\t")
            elements.reverse()

            # inserting the values into the cell
            for cell in row:
                cell.value = elements.pop()

        # saving the file
        wb.save(self.__path + "dataset.xlsx")

    def singleColumnSettings(self, settingsList, uniqueColumn):
        # loading the file
        wb = load_workbook(self.__path + "dataset.xlsx")
        ws = wb.active
        ws.active = 1

        # transferring the cost centre settings
        colNo = uniqueColumn
        rowLimit = len(settingsList) + 1
        ccList = list(settingsList)

        for row in ws.iter_rows(
            min_row=2, min_col=colNo, max_col=colNo, max_row=rowLimit
        ):
            for cell in row:
                settingsItem = ccList.pop()
                cell.value = settingsItem

        # saving the file
        wb.save(self.__path + "dataset.xlsx")

