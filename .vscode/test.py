from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws["A1"] = "Column A"
ws["B1"] = "Column B"
ws["C1"] = "Column C"

ws["A2"] = "Cell A1"
ws["B2"] = "Cell B2"
ws["C2"] = "Cell B3"

ws["A3"] = "Cell A2"
ws["B3"] = "Cell B2"
ws["C3"] = "Cell B3"

ws["A4"] = "Cell A4"
ws["B4"] = "Cell B4"

ws["A5"] = "Cell A5"


# looping through the column
myList = []
for cell in ws["A"]:
    x = cell.value
    myList.append(x)
    print(myList)


print(type(myList))

for item in myList:
    ws.append(myList)


wb.save("test.xlsx")

